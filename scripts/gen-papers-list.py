#!/usr/bin/env python3
"""Print the CS858 paper list as a verbatim markdown table.

Reads vendor/cs858-wiki/docs/CS858-F26-papers-stripped.xlsx (sheet
UpdatedList) and writes a pipe table to stdout — one row per paper, with
essential and extra readings joined by a semicolon. No links, labels, or
site-specific markup are added.

Run from the site root::

    uv run --with openpyxl python3 scripts/gen-papers-list.py

Paste the output into the body of src/content/pages/cs858-papers-list.md.
"""

from __future__ import annotations

from pathlib import Path
from typing import cast

from openpyxl import load_workbook

XLSX = Path("vendor/cs858-wiki/docs/CS858-F26-papers-stripped.xlsx")
SHEET = "UpdatedList"

NUMBER_COL = 1
TITLE_COL = 2
THEME_COL = 3
TOPIC_COL = 4
ESSENTIAL_COL = 5


def cell_item(row: tuple, col: int) -> str:
    """Return the cell as a markdown link if it has a hyperlink, else plain text."""
    cell = row[col - 1]
    text = " ".join(str(cell.value).split()) if cell.value is not None else ""
    if not text:
        return ""
    href = cell.hyperlink.target if cell.hyperlink else None
    if href:
        safe = text.replace("[", "\\[").replace("]", "\\]").replace("|", "\\|")
        return f"[{safe}]({href})"
    return text.replace("|", "\\|")


def esc(text: str) -> str:
    return text.replace("|", "\\|")


def main() -> None:
    wb = load_workbook(XLSX)
    ws = wb[SHEET]
    max_row = cast(int, ws.max_row)
    max_col = cast(int, ws.max_column)

    parts: list[dict] = []
    current_part: dict | None = None
    current_paper: dict | None = None
    current_theme = ""

    for row in ws.iter_rows(min_row=2, max_row=max_row):
        num_raw = row[NUMBER_COL - 1].value
        theme_raw = cell_item(row, THEME_COL)
        if theme_raw:
            current_theme = theme_raw

        if isinstance(num_raw, str) and num_raw.startswith("Part"):
            current_part = {"title": num_raw, "papers": []}
            parts.append(current_part)
            current_paper = None
            continue

        if isinstance(num_raw, int):
            essential = cell_item(row, ESSENTIAL_COL)
            current_paper = {
                "number": num_raw,
                "title": cell_item(row, TITLE_COL),
                "theme": current_theme,
                "topic": cell_item(row, TOPIC_COL),
                "essential": [essential] if essential else [],
            }
            if current_part is not None:
                current_part["papers"].append(current_paper)
            continue

        if num_raw is None and current_paper is not None:
            essential = cell_item(row, ESSENTIAL_COL)
            if essential:
                current_paper["essential"].append(essential)

    blocks: list[str] = []
    for part in parts:
        title = part["title"]
        words = title.split(" ", 2)
        heading = (
            f"## Part {words[1]}: {words[2]}"
            if len(words) == 3 and words[0] == "Part"
            else f"## {title}"
        )
        blocks.append(heading)

        header = "| # | Paper | Theme | Topic | Essential readings |"
        sep = "|---|-------|-------|-------|-------------------|"
        rows = [header, sep]
        for p in sorted(part["papers"], key=lambda x: x["number"]):
            essential = "; ".join(p["essential"])
            rows.append(
                f"| {p['number']} | {p['title']} | {esc(p['theme'])} |"
                f" {esc(p['topic'])} | {essential} |"
            )
        blocks.append("\n".join(rows))

    print("\n\n".join(blocks))


if __name__ == "__main__":
    main()
